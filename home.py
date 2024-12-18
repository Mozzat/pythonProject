import openpyxl
import util
import SQLManager
from openpyxl.styles import Font, Color, Border, Side, PatternFill, Alignment
from tkinter import filedialog
from datetime import datetime
from openpyxl.utils import get_column_letter

filePath = ''

def marchMoreProductsData(data):

    # 原料总成本   
    y_cost = 0
    # 原料实际领用量
    y_costVolume = 0
    # 辅料实际领用量
    f_costVolume = 0
    f_cost = 0
    totalCost = 0
    y_bgs = 0
    f_bgs = 0

    ggZsyL = 0

    count = len(data)
    for index in range(count):
        item = data[index]
        itemCount = len(item) - 1
        for j in range(itemCount):
            item1 = item[j]
            if j == 0:
                #计算报工数
                firstData = item1
                f_bgs += int(firstData[4])
                if util.jundgeIsLowPriceProduct(firstData[2]) == False:
                    y_bgs += int(firstData[4])
                    w = util.getResultWeight(firstData[2])
                    ggZsyL += float(w) * int(firstData[4])
                #计算总成本
                lastData = item[itemCount]
                totalCost += float(lastData[10])
            # 判断是不是原料 数组大于0 表示查询产品库查询到了数据
            # productArr = SQLManager.queryProductByProductCode(item1[7])
            if item1[6] == "原料":
                cost = float(item1[10])
                ck = item1[14]
                y_cost += cost
                sjlql = item1[9]
                sjlqlS = str(sjlql).lower()
                if sjlqlS.find("g") != -1:
                    sjlqlNum = util.getProductQualityByName(str(item1[9]))
                    sjlql = sjlqlNum / 1000
                else:
                    #特殊情况 产品是成品 用产品的容量*领取量
                    weightTemp = util.getProductWeightByName(item1[8])
                    if str(weightTemp).lower().find("g") != -1:
                        weightNumTemp = util.getProductQualityByName(weightTemp)
                        sjlql = (weightNumTemp / 1000.0) * sjlql
                y_costVolume += float(sjlql)
            else:
                f_cost += float(item1[10])
                f_costVolume += float(item1[9])
            
    # 计算原料摊销 原料摊销金额 辅料摊销金额 单品成本 出品率
    #看是不是余料仓，如果是，原料的总成本为0，但实际领用量保持不变
    # 1.总成本：重复的项目总成本相加，求和
    # 2.实际领用量：总成本/物料出库价，求和
     # 3.原料摊销KG=原料实际领用总量/（品的报工数*货品规格的合计）*品的报工数*规格，记得锁定
    # 4.原料摊销金额=原料总成本/（品的报工数*货品规格）*品的报工数*规格，记得锁定
    # 5.辅料摊销金额=辅料总成本/品的总报工数*单品的报工数，记得锁定
    # 单片成本=（原料摊销金额+辅料摊销金额）/各单品的报工数
    # 出品率=单品报工数*规格/原料摊销KG（记得换算）
    if ggZsyL == 0:
        ggZsyL = 1

    for index in range(count):
        item = data[index]
        itemCount = len(item) - 1
        lastData = item[itemCount]
        hasYL = False
        for j in range(itemCount):
            item1 = item[j]
            if j == 0:
                weightNum = util.getResultWeight(item1[2])
            # 判断是不是原料 数组大于0 表示查询产品库查询到了数据
            # productArr = SQLManager.queryProductByProductCode()
            if item1[6] == "原料":
                hasYL = True
                bgs = item1[4]
                y_bgs = 1 if y_bgs == 0 else y_bgs
                f_bgs = 1 if f_bgs == 0 else f_bgs
                yltx = (y_costVolume / ggZsyL) * bgs * weightNum
                yltxje = (y_cost / ggZsyL)* bgs * weightNum
                fltxje = (f_cost / f_bgs) * bgs
                dpcost = (yltxje + fltxje) / bgs
                if yltx == 0:
                    yltx = 1
                chpr = (bgs * weightNum) / yltx
                ck = item1[14]
                # 看是不是余料仓 如果是余料仓 成本是0
                if ck == '余料仓' or util.jundgeIsLowPriceProduct(item1[2]):
                    yltx = 0
                    yltxje = 0
                    dpcost = fltxje / bgs
                    chpr = 0
                print('原料摊销KG：' + str(yltx), '原料摊销金额：' + str(yltxje), '辅料摊销金额：' + str(fltxje), '单片成本：' + str(dpcost), '出品率：' + str(chpr))
                lastData1 = list(lastData)
                lastData1[2] = "" if yltx == 0 else yltx
                lastData1[3] = "" if yltxje == 0 else yltxje
                lastData1[4] = "" if fltxje == 0 else fltxje
                lastData1[5] = "" if dpcost == 0 else '￥' + str(dpcost)
                lastData1[6] = "" if chpr == 0 else str(round(chpr * 100,2)) + '%'
                lastData1[9] = str(y_costVolume + f_costVolume) 
                lastData1[10] = str(round(totalCost,2)) 
                lastData2 = tuple(lastData1)
                item[itemCount] = lastData2
                continue
        # 23AFA2963
        if hasYL == False:
            bgs = item1[4] 
            y_bgs = 1 if y_bgs == 0 else y_bgs
            f_bgs = 1 if f_bgs == 0 else f_bgs
            yltx = (y_costVolume / ggZsyL) * bgs * weightNum
            yltxje = (y_cost / ggZsyL)* bgs * weightNum
            fltxje = (f_cost / f_bgs) * bgs
            dpcost = (yltxje + fltxje) / bgs
            if yltx == 0:
                    yltx = 1
            chpr = (bgs * weightNum) / yltx
            ck = item1[14]
            # 看是不是余料仓 如果是余料仓 成本是0
            if ck == '余料仓' or util.jundgeIsLowPriceProduct(item1[2]):
                yltx = 0
                yltxje = 0
                dpcost = fltxje / f_bgs
                chpr = 0
            print('原料摊销KG：' + str(yltx), '原料摊销金额：' + str(yltxje), '辅料摊销金额：' + str(fltxje), '单片成本：' + str(dpcost), '出品率：' + str(chpr))
            lastData1 = list(lastData)
            lastData1[2] = "" if yltx == 0 else yltx
            lastData1[3] = "" if yltxje == 0 else yltxje
            lastData1[4] = "" if fltxje == 0 else fltxje
            lastData1[5] = "" if dpcost == 0 else '￥' + str(dpcost)
            lastData1[6] = "" if chpr == 0 else str(round(chpr * 100,2)) + '%'
            lastData1[9] = str(round(y_costVolume + f_costVolume,2)) 
            lastData1[10] = str(round(totalCost,2)) 
            lastData2 = tuple(lastData1)
            item[itemCount] = lastData2
            

    firstCount = len(data[0]) - 1
    tempArr = []
    for index in range(firstCount):
        firstModel = data[0][index]
        totalLql = firstModel[9]
        totalLqlS = str(totalLql).lower()
        if totalLqlS.find("g") != -1:
            totalLqlNum = util.getProductQualityByName(str(firstModel[9]))
            totalLql = totalLqlNum / 1000
        total = firstModel[10]
        for j in range(count):
            if j == 0:
                continue
            else:
                item = data[j]
                itemCount = len(item) - 1
                for z in range(itemCount):
                    item1 = item[z]
                    if item1[7] == firstModel[7]:
                        sjlql = item1[9]
                        sjlqlS = str(sjlql).lower()

                        if sjlqlS.find("g") != -1:
                            sjlqlNum = util.getProductQualityByName(str(item1[9]))
                            sjlql = sjlqlNum / 1000
                        else:
                            #特殊情况 产品是成品 用产品的容量*领取量
                            weightTemp = util.getProductWeightByName(item1[8])
                            if str(weightTemp).lower().find("g") != -1:
                                weightNumTemp = util.getProductQualityByName(weightTemp)
                                sjlql = (weightNumTemp / 1000.0) * sjlql
                        total += item1[10]
                        totalLql += sjlql
                    else:
                        continue
        arr = list(firstModel)
        arr[9] = str(round(totalLql,2))
        arr[10] = str(round(total,2))
        firstModel1 = tuple(arr)
        tempArr.append(firstModel1)
    tempArr.append(data[0][firstCount])
    data[0] = tempArr
    return data

def queryFileReadData(filePath):
    filePath = filePath
    # exclePath = '/Users/mozzat/Desktop/pythonProject/领料汇总表.xlsx'
    exclePath = filePath
    wb = openpyxl.load_workbook(exclePath)
    sheet = wb.active

    print('==============读取数据==============')
    totalArr = []
    #  分组处理数据
    #  ('任务单', '货品编号', '货品名称', '任务数', '报工数', '待报工数', '物料类型', '物料编号', '物料名称', '实际领用量', '总成本', '未申请数量', '领料使用量', '领料剩余量', '物料出库仓', '物料出库价', '领料剩余价值', '生产订单', '生产成本单价', '分类', '退料数量')
    excleDic = {}
    lastKey = ''
    for row in sheet.iter_rows(min_row=1,max_row=sheet.max_row, values_only=True):
        firstName = row[0]
        if firstName == '任务单':
            continue
        if firstName != None:
            lastKey = firstName
            excleArr = excleDic.get(lastKey)
            if excleArr == None:
                excleArr = []
                excleArr.append(row)
                excleDic[lastKey] = excleArr
            else:
                excleArr.append(row)
                excleDic[lastKey] = excleArr
        else:
            excleArr = excleDic.get(lastKey)
            excleArr.append(row)
            excleDic[lastKey] = excleArr

    excleKeyArr = excleDic.keys()
    for key in excleKeyArr:
        excleArr = excleDic[key]
        totalArr.append(excleArr)
    
    # totalArr数组通过同一个生产订单再次合并
    resultArr = []
    resultDic = {}

    for item in totalArr:
        key = item[0][17]
        tempArr = resultDic.get(key)
        if tempArr == None:
            tempArr = []
            tempArr.append(item)
            resultDic[key] = tempArr
        else:
            tempArr.append(item)
            resultDic[key] = tempArr

    keyArr = resultDic.keys()
    for key in keyArr:
        resultArr.append(resultDic[key])        

    resultCount = len(resultArr)
    for index in range(resultCount):
        tempArr = resultArr[index]
        if len(tempArr) > 1:
            # 对临时数据处理 T/B 产品不放在第一位
            normalArr = []
            tbArr = []
            if index == 10:
                print(index)
            for tempItem in tempArr:
                if util.jundgeIsLowPriceProduct(tempItem[0][2]):
                    tbArr.append(tempItem)
                else:
                    normalArr.append(tempItem)
            normalArr1 = sorted(normalArr, key=len, reverse=True)
            tempArr1 = normalArr1 + tbArr
            resultArr[index] = tempArr1   


    # 合并产品模型 并且合并计算各个产品原料的的总成本


    #  整合数据
    # 需要计算的数据 原料摊销 原料摊销金额 辅料摊销金额 单品成本 出品率
    # 计算公式：
    ## 单品 
    # 单片成本 = 品的总成本/品的总报工数
    # 出品率 = 总报工数 * 货品规格 / 总实际领用量
    # 重量从产品名称中取值比较准确 要考虑到 克 和千克 

#    日期	任务单	货品编号	货品名称	任务数	报工数	待报工数	
#   原料摊销	原料摊销金额	辅料摊销金额	单品成本	出品率	物料类型	
# 物料编号	物料名称	实际领用量	总成本	未申请数量	领料使用量	领料剩余量	
# 物料出库仓	物料出库价	领料剩余价值	生产订单	生产成本单价	分类	退料数量
    firstTurbe = ("日期","任务单","货品编号","货品名称","任务数","报工数","待报工数","原料摊销","原料摊销金额","辅料摊销金额","单品成本","出品率","物料类型","物料编号","物料名称","实际领用量","总成本","未申请数量","领料使用量","领料剩余量","物料出库仓","物料出库价","领料剩余价值","生产订单","生产成本单价","分类","退料数量")
    data1 = []
    data1.append(firstTurbe)


    
    for item in resultArr:
        resultData = marchMoreProductsData(item)
        
        #创建导出的数据
        length = len(resultData)
        length1 = len(resultData[0]) -1
        length2 = length if length > length1 else length1

        firstItem = resultData[0]
        firstData = resultData[0][0]
        lastItem = firstItem[length1]
        for index in range(length2):
            if index == 0: 
                turbe = ("",firstData[0],firstData[1],firstData[2],firstData[3],firstData[4],firstData[5],lastItem[2],lastItem[3],lastItem[4],lastItem[5],lastItem[6],firstData[6],firstData[7],firstData[8],firstData[9],firstData[10],firstData[11],firstData[12],firstData[13],firstData[14],firstData[15],firstData[16],firstData[17],firstData[18],firstData[19],firstData[20])
                data1.append(turbe)
            else:
                trube = []
                tempTrube1 = []
                tempTrube2 = []
                lastTrube1 = []
                hasData = 0
                hasData1 = 0
                if index < length1:
                    hasData = 1
                    tempTrube1 = list(firstItem[index])
                if index < length:
                    hasData1 = 1
                    tempTrube2 = list(resultData[index][0])
                    j = len(resultData[index]) - 1
                    lastTrube1 = list(resultData[index][j])
                # 数据都有值
                if hasData == 1 and hasData1 == 1:
                    trube1 = ("",firstData[0],tempTrube2[1],tempTrube2[2],tempTrube2[3],tempTrube2[4],tempTrube2[5],lastTrube1[2],lastTrube1[3],lastTrube1[4],lastTrube1[5],lastTrube1[6],tempTrube1[6],tempTrube1[7],tempTrube1[8],tempTrube1[9],tempTrube1[10],tempTrube1[11],tempTrube1[12],tempTrube1[13],tempTrube1[14],tempTrube1[15],tempTrube1[16],tempTrube1[17],tempTrube1[18],tempTrube1[19],tempTrube1[20])
                    trube = list(trube1)
                elif hasData == 1 and hasData1 == 0:
                    trube1 = ("",firstData[0],"","","","","","","","","","",tempTrube1[6],tempTrube1[7],tempTrube1[8],tempTrube1[9],tempTrube1[10],tempTrube1[11],tempTrube1[12],tempTrube1[13],tempTrube1[14],tempTrube1[15],tempTrube1[16],tempTrube1[17],tempTrube1[18],tempTrube1[19],tempTrube1[20])
                    trube = list(trube1)
                elif hasData == 0 and hasData1 == 1:
                    trube1 = ("",firstData[0],tempTrube2[1],tempTrube2[2],tempTrube2[3],tempTrube2[4],tempTrube2[5],lastTrube1[2],lastTrube1[3],lastTrube1[4],lastTrube1[5],lastTrube1[6],"","","","","","","","","","","","","","",)
                    trube = list(trube1)
                data1.append(tuple(trube))
        costTurbe = ("",firstData[0],"","","","","","","","","","","","","",lastItem[9],lastItem[10],"","","","","","","","","",)
        data1.append(costTurbe)
    exportExcleFile(data1)

def exportExcleFile(data1):

    wb = openpyxl.Workbook()
    sheet = wb.active
    # 设置单元格的字体
    # font = Font(name='Arial', size=14, bold=True, italic=True, color="FF0000")
    # sheet['A1'] = font
    # 设置单元格的背景颜色
    # sheet['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 设置单元格的边框
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    for row in data1:
        sheet.append(row)
    

    # 设置excle样式
    for col in sheet.columns: 
        for cell in col:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Arial', size=12)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # 自动调整单元格的宽度
    for col in sheet.columns:
        max_length = 20
        column = col[0].column  # 获取列号
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    # 合并单元格
    column = 'B'
    cells = sheet[column]
    # 初始化变量
    start_row = 1
    end_row = 1

    # 遍历 A 列的所有单元格
    for i in range(1, len(cells)):
        # 如果当前单元格的值与前一个单元格的值相同
        if cells[i].value == cells[i-1].value:
            # 更新结束行号
            end_row = i + 1
        else:
            # 如果当前单元格的值与前一个单元格的值不同
            # 合并从开始行到结束行的单元格
            sheet.merge_cells(column + str(start_row) + ':' + column + str(end_row))
            # 更新开始行号
            start_row = i + 1
            end_row = i + 1

    # 合并最后一组相连的单元格
    sheet.merge_cells(column + str(start_row) + ':' + column + str(end_row))

    # 获取当前日期和时间
    now = datetime.now()

    # 格式化日期和时间
    formatted_now = now.strftime("%Y-%m-%d %H:%M:%S")
    # 设置默认文件名和保存类型
    default_file = '数据'+ formatted_now + '.xlsx'
 
    # 弹出保存文件对话框
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], initialfile=default_file)
    if save_path:
        # 保存Excel文件
        wb.save(save_path)

# queryFileReadData('/Users/mozzat/Desktop/pythonProject/工作簿1.xlsx')